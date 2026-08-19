"""입력 형식 분기 — PDF · 엑셀 · 이미지를 같은 `ParsedDocument` 로 만든다.

진단 자료가 PDF 만 오는 것은 아니다. 계측 데이터는 **엑셀**로 오고, 설비 명판과
현장 사진은 **이미지**로 온다. 뒤쪽 파이프라인(업종 분류 · 적재 게이트 · 온톨로지 ·
위키 생성)은 전부 `ParsedDocument` 하나만 보므로, 형식별 차이는 여기서 끝낸다.

`parse.py`(PDF)와 나란히 두고 그 위에 얹지 않는다. 형식이 늘 때마다 PDF 파서를
고치면 언젠가 PDF 파싱이 엑셀 때문에 깨진다.

| 형식 | 무엇이 되는가 | 한계 |
|---|---|---|
| `.pdf` | 글·표·그림 4채널 (`parse.parse_pdf`) | 스캔본은 텍스트 레이어가 없어 빈 결과 |
| `.xlsx` `.xlsm` `.csv` | 시트 = 쪽, 빈 줄로 끊어 표 블록 | 수식은 **저장된 계산값**을 읽는다 |
| `.png` `.jpg` … | 그림 채널 + (OCR 있으면) 글 | **OCR 이 없으면 글자를 못 읽는다** |

이미지에서 글자를 못 읽는 것은 조용히 넘기지 않는다. 빈 결과를 주면 사용자는
'분석했는데 아무것도 없다'고 읽고, 그건 '읽지 못했다'와 전혀 다른 뜻이다.
"""

from __future__ import annotations

import csv
import hashlib
import os
from typing import Any

from .parse import (
    ImageBlock,
    ParseError,
    ParsedDocument,
    TableBlock,
    TextBlock,
    _clean,
    _is_numeric,
    parse_pdf,
)

#: 받아들이는 확장자 → 형식. 닫힌 집합이다 — 여기 없는 확장자는 업로드 단계에서 막힌다.
SUFFIX_KIND: dict[str, str] = {
    ".pdf": "pdf",
    ".xlsx": "sheet", ".xlsm": "sheet", ".xltx": "sheet", ".csv": "sheet", ".tsv": "sheet",
    ".png": "image", ".jpg": "image", ".jpeg": "image", ".webp": "image",
    ".bmp": "image", ".tif": "image", ".tiff": "image",
}

KINDS: tuple[str, ...] = ("pdf", "sheet", "image")

#: 시트를 표 블록으로 끊는 기준. 빈 줄이 이만큼 이어지면 다른 표로 본다.
BLANK_ROWS_SPLIT = 1

#: 이보다 긴 시트는 뒤를 자른다. 계측 로그는 수만 행이 오는데, 전부 넣으면 검색
#: 레코드가 로그로 뒤덮여 정작 진단 내용이 묻힌다.
MAX_SHEET_ROWS = 2000


def kind_of(path: str) -> str:
    """확장자로 형식을 정한다. 모르는 확장자는 받지 않는다."""
    suffix = os.path.splitext(str(path).lower())[1]
    kind = SUFFIX_KIND.get(suffix)
    if kind is None:
        raise ParseError(
            f"지원하지 않는 형식이다: {suffix or '확장자 없음'} "
            f"(가능: {', '.join(sorted(SUFFIX_KIND))})")
    return kind


def parse_document(path: str, *, extract_images: bool = True) -> ParsedDocument:
    """형식을 가려 파싱한다. 뒤쪽 파이프라인은 이 함수만 부르면 된다."""
    kind = kind_of(path)
    try:
        if kind == "pdf":
            return parse_pdf(path, extract_images=extract_images)
        if kind == "sheet":
            return parse_spreadsheet(path)
        return parse_image(path)
    except ParseError as exc:
        # 사유는 그대로 두고 **어느 파일인지**만 앞에 붙인다. 여러 건을 돌릴 때
        # '엑셀을 열 수 없다' 만으로는 어느 것이 문제인지 알 수 없다.
        name = os.path.basename(path)
        raise ParseError(str(exc) if name in str(exc) else f"{name}: {exc}") from exc
    except Exception as exc:  # noqa: BLE001 - 사용자가 올린 파일이다. 스택트레이스가 아니라
        # '무엇이 잘못됐는지'를 돌려준다. 확장자는 맞는데 내용이 깨진 경우가 대부분이다.
        raise ParseError(
            f"{os.path.basename(path)} 를 읽을 수 없다 ({kind}): {exc}") from exc


def _hash(path: str) -> str:
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()[:16]


# --------------------------------------------------------------------------- #
# 엑셀 · CSV
# --------------------------------------------------------------------------- #
def parse_spreadsheet(path: str) -> ParsedDocument:
    """시트를 쪽으로, 빈 줄로 끊은 덩어리를 표로 읽는다.

    한 시트에 표 하나만 있다고 보면 안 된다. 진단 자료의 엑셀은 제목 몇 줄, 표 하나,
    빈 줄, 또 표 하나가 이어지는 형태가 흔하다. 통째로 한 표로 만들면 헤더가
    엉뚱한 줄이 되어 모든 열 이름이 틀어진다.

    수식은 **저장된 계산값**으로 읽는다(`data_only=True`). 수식 문자열(`=B2*C2`)이
    값으로 들어가면 수치 검산이 통째로 무의미해진다. 엑셀이 한 번도 열린 적 없어
    캐시가 비어 있으면 값이 None 이 되는데, 그건 경고로 남긴다.
    """
    doc = ParsedDocument(filename=os.path.basename(path), doc_hash=_hash(path), n_pages=0)
    suffix = os.path.splitext(path.lower())[1]

    if suffix in (".csv", ".tsv"):
        sheets = [(os.path.basename(path), _read_csv(path, "\t" if suffix == ".tsv" else ","))]
    else:
        sheets = _read_workbook(path, doc)

    doc.n_pages = len(sheets)
    if not sheets:
        doc.warnings.append("읽을 수 있는 시트가 없다")
        return doc

    for page, (name, rows) in enumerate(sheets, start=1):
        if len(rows) > MAX_SHEET_ROWS:
            doc.warnings.append(
                f"시트 '{name}' 이 {len(rows):,}행이라 앞 {MAX_SHEET_ROWS:,}행만 읽었다")
            rows = rows[:MAX_SHEET_ROWS]

        tidx = 0
        for block in _split_blocks(rows):
            # 한 칸짜리 줄만 모인 덩어리는 표가 아니라 제목·메모다.
            if max((len(r) for r in block), default=0) < 2:
                text = "\n".join(" ".join(r) for r in block).strip()
                if text:
                    doc.text_blocks.append(
                        TextBlock(page=page, idx=len(doc.text_blocks), text=f"{name}\n{text}"))
                continue
            block = _drop_empty_columns(block)
            header, body = block[0], block[1:]
            if not body:
                continue
            width = max(len(r) for r in block)
            header = (header + [""] * width)[:width]
            body = [(r + [""] * width)[:width] for r in body]
            n_num = sum(1 for r in body for c in r if _is_numeric(c))
            doc.tables.append(TableBlock(page=page, idx=tidx, header=header, rows=body,
                                         n_numeric_cells=n_num, caption=name))
            tidx += 1

        if tidx == 0 and not any(b.page == page for b in doc.text_blocks):
            doc.warnings.append(f"시트 '{name}' 에서 표를 찾지 못했다")

    return doc


def _read_workbook(path: str, doc: ParsedDocument) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 설치 안내 경로
        raise ParseError("openpyxl 이 필요하다 (uv pip install openpyxl)") from exc

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - 깨진 파일을 빈 결과로 넘기지 않는다
        raise ParseError(f"엑셀을 열 수 없다: {exc}") from exc

    out: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for raw_row in ws.iter_rows(values_only=True):
            cells = [_clean(c) for c in raw_row]
            while cells and not cells[-1]:
                cells.pop()
            rows.append(cells)
        out.append((ws.title, rows))
    wb.close()

    missing = _formulas_without_values(path)
    if missing:
        doc.warnings.append(
            f"수식 {missing}칸의 계산값이 저장돼 있지 않아 빈 값으로 읽었다 "
            "(엑셀에서 한 번 열어 저장하면 채워진다)")
    return out


def _formulas_without_values(path: str) -> int:
    """수식은 있는데 계산값이 없는 칸 수.

    '빈 칸이 많다'로 짐작하면 서식만 잡힌 성긴 시트가 전부 경고를 받는다. 수식이
    실제로 있는지 확인하고서 센다 — 여기서 나오는 값이 0 이 아니면 그 파일의 수치는
    믿을 수 없다.
    """
    try:
        from openpyxl import load_workbook

        formulas = load_workbook(path, data_only=False, read_only=True)
        values = load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 - 확인에 실패해도 파싱은 계속한다
        return 0
    count = 0
    try:
        for fs, vs in zip(formulas.worksheets, values.worksheets):
            for frow, vrow in zip(fs.iter_rows(values_only=True), vs.iter_rows(values_only=True)):
                for f, v in zip(frow, vrow):
                    if isinstance(f, str) and f.startswith("=") and v is None:
                        count += 1
    finally:
        formulas.close()
        values.close()
    return count


def _read_csv(path: str, delimiter: str) -> list[list[str]]:
    # 한국어 CSV 는 cp949 로 저장돼 오는 일이 흔하다. utf-8 로만 읽으면 통째로 깨진다.
    for encoding in ("utf-8-sig", "cp949", "utf-8"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                return [[_clean(c) for c in row] for row in csv.reader(f, delimiter=delimiter)]
        except UnicodeDecodeError:
            continue
    raise ParseError("CSV 인코딩을 알 수 없다 (utf-8 · cp949 로 읽어 보았다)")


def _drop_empty_columns(block: list[list[str]]) -> list[list[str]]:
    """덩어리 안에서 처음부터 끝까지 빈 열은 버린다.

    엑셀 서식에서 여백으로 비워 둔 열이 그대로 남으면 헤더가 `['', '', '항목']` 이 되어
    열 이름으로 표를 찾는 추출기가 전부 헛돈다.
    """
    width = max((len(r) for r in block), default=0)
    keep = [i for i in range(width) if any(i < len(r) and r[i] for r in block)]
    if len(keep) == width:
        return block
    return [[r[i] if i < len(r) else "" for i in keep] for r in block]


def _split_blocks(rows: list[list[str]]) -> list[list[list[str]]]:
    """빈 줄로 끊어 덩어리를 만든다."""
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    blanks = 0
    for row in rows:
        if not any(row):
            blanks += 1
            if current and blanks >= BLANK_ROWS_SPLIT:
                blocks.append(current)
                current = []
            continue
        blanks = 0
        current.append(row)
    if current:
        blocks.append(current)
    return blocks


# --------------------------------------------------------------------------- #
# 이미지
# --------------------------------------------------------------------------- #
def ocr_ready() -> dict[str, Any]:
    """이미지에서 글자를 읽을 수 있는 상태인가.

    화면이 **업로드하기 전에** 알아야 한다. 올리고 나서 '아무것도 없음'을 보면
    파일이 비어 있는 것인지 도구가 없는 것인지 알 수 없다.
    """
    try:
        import pytesseract
    except ImportError:
        return {"ok": False, "reason": "OCR 도구(pytesseract)가 설치되지 않았다",
                "hint": "uv pip install pytesseract 후 tesseract 본체와 한국어 데이터가 필요하다\n"
                        "  Ubuntu: sudo apt install tesseract-ocr tesseract-ocr-kor\n"
                        "  macOS : brew install tesseract tesseract-lang"}
    try:
        version = str(pytesseract.get_tesseract_version())
    except Exception as exc:  # noqa: BLE001 - 실행 파일이 없거나 경로가 틀린 경우
        return {"ok": False, "reason": f"tesseract 본체를 실행할 수 없다: {exc}",
                "hint": "Ubuntu: sudo apt install tesseract-ocr tesseract-ocr-kor"}
    langs: list[str] = []
    try:
        langs = list(pytesseract.get_languages(config=""))
    except Exception:  # noqa: BLE001 - 언어 목록은 없어도 OCR 자체는 된다
        pass
    return {"ok": True, "reason": "", "hint": "", "version": version, "languages": langs}


def parse_image(path: str) -> ParsedDocument:
    """이미지 1장 = 1쪽. OCR 이 있으면 글도 읽고, 없으면 **없다고 말한다.**"""
    try:
        from PIL import Image
    except ImportError as exc:  # pragma: no cover - 설치 안내 경로
        raise ParseError("Pillow 가 필요하다 (uv pip install pillow)") from exc

    doc = ParsedDocument(filename=os.path.basename(path), doc_hash=_hash(path), n_pages=1)
    try:
        with Image.open(path) as im:
            width, height = im.size
            im.load()
            frame = im.convert("RGB") if im.mode not in ("RGB", "L") else im.copy()
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f"이미지를 열 수 없다: {exc}") from exc

    ready = ocr_ready()
    text = ""
    if ready["ok"]:
        import pytesseract

        # 한국어 우선, 없으면 설치된 것으로 시도한다. 언어를 지정하지 않으면
        # 영어로 읽어 한글이 통째로 깨진다.
        langs = ready.get("languages") or []
        lang = "kor+eng" if "kor" in langs else ("eng" if "eng" in langs else None)
        try:
            text = pytesseract.image_to_string(frame, lang=lang) if lang else \
                pytesseract.image_to_string(frame)
        except Exception as exc:  # noqa: BLE001
            doc.warnings.append(f"OCR 실패: {exc}")
        if "kor" not in langs:
            doc.warnings.append(
                "한국어 OCR 데이터(tesseract-ocr-kor)가 없어 한글은 읽지 못한다")
    else:
        doc.warnings.append(f"이미지에서 글자를 읽지 못했다 — {ready['reason']}")

    if text.strip():
        for idx, chunk in enumerate(p for p in text.split("\n\n") if p.strip()):
            doc.text_blocks.append(TextBlock(page=1, idx=idx, text=chunk.strip()))

    doc.images.append(ImageBlock(
        page=1, idx=0, width=width, height=height,
        kind=_image_kind(width, height, text),
        nearby_caption=os.path.splitext(os.path.basename(path))[0]))
    return doc


def _image_kind(width: int, height: int, text: str) -> str:
    """종류 추정. 확정이 아니라 제안이다 — 사람이 뒤집을 수 있다."""
    flat = (text or "")
    if any(k in flat for k in ("명판", "MODEL", "형식", "정격", "SERIAL")):
        return "drawing"
    ratio = width / max(height, 1)
    if ratio > 3.0 or ratio < 0.33:
        return "drawing"
    return "photo"


# --------------------------------------------------------------------------- #
def readiness() -> dict[str, Any]:
    """형식별 준비 상태. `/api/kb/health` 가 그대로 화면에 내보낸다."""
    try:
        import pdfplumber  # noqa: F401

        pdf = {"ok": True, "reason": "", "hint": ""}
    except ImportError:
        pdf = {"ok": False, "reason": "pdfplumber 가 설치되지 않았다",
               "hint": "uv pip install pdfplumber"}
    try:
        import openpyxl  # noqa: F401

        sheet = {"ok": True, "reason": "", "hint": ""}
    except ImportError:
        sheet = {"ok": False, "reason": "openpyxl 이 설치되지 않았다",
                 "hint": "uv pip install openpyxl"}
    return {
        "pdf": pdf,
        "sheet": sheet,
        # 이미지는 두 단이다: 파일을 여는 것(Pillow)과 글자를 읽는 것(OCR).
        # 앞은 되고 뒤는 안 되는 상태가 정상적으로 있을 수 있어 따로 알린다.
        "image": ocr_ready(),
        "suffixes": sorted(SUFFIX_KIND),
    }
