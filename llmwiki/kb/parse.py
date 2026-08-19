"""문서 4채널 분해 — 글 · 표 · 그림 · 엑셀.

`llmwiki/compliance/docparse.py` 와 목적이 다르다. 저쪽은 규정 문서를 **하나의
평문 좌표계**로 펴서 근거 스팬의 기준을 만든다(조문 앵커가 목적). 여기는 진단
보고서를 **채널로 갈라** 표의 격자를 살린다(수치의 출처가 목적).

통짜 텍스트로 만들면 표의 행 구조가 깨진다. 에너지진단 보고서 32면에서 실측한
결과, 표 데이터 행 95개 중 행 구조(셀 순서·인접)가 보존된 것은 21개(22%)뿐이었다.
숫자 자체는 텍스트 스트림 어딘가에 남지만 **"어느 행 어느 열의 값인가"** 가
사라져서 ``18대 | 22kW | 25.7kW | 7,200h | 80% | 2,664,576kWh`` 가 의미를 잃는다.
진단 보고서는 수치가 전부 표에 있으므로 그 상태로는 문서의 절반만 검색된다.

| 채널  | 무엇             | 앵커                            | 쓰임                     |
|-------|------------------|---------------------------------|--------------------------|
| text  | 문단·항목 서술   | page                            | 문제점·개선방안 검색     |
| table | 셀 그리드        | page + table_idx + (row, col)   | **모든 수치의 출처**     |
| image | 사진·도면·차트   | page + image_idx                | 증적, 명판 OCR(v0.2)     |
| excel | 표 전체를 시트로 | sheet = p{page}_t{idx}          | 감리 제출, 재계산        |

앵커는 ``(문서해시, 페이지, 표인덱스, 행, 열)`` 이다. bbox 는 속성으로만 들고
ID 에 넣지 않는다 — 파서를 고치면 좌표가 밀려 그래프가 통째로 끊기기 때문이다
(`ontology.py` 의 단일 실패 지점).
"""

from __future__ import annotations

import hashlib
import io
import os
import re
from dataclasses import asdict, dataclass, field
from typing import Any

NUMERIC = re.compile(r"-?\d[\d,]*\.?\d*")

#: 로고·머리말 같은 잡동사니 이미지를 걸러내는 최소 크기(px)
MIN_IMAGE_PX = 80

#: 채널의 닫힌 집합. 여기 없는 채널은 청크에 나올 수 없다.
CHANNELS: tuple[str, ...] = ("text", "table", "image", "excel")

#: 이미지 종류 추정값의 닫힌 집합. 확정이 아니라 제안이다 — 사람이 뒤집을 수 있다.
IMAGE_KINDS: tuple[str, ...] = ("photo", "drawing", "chart", "logo", "unknown")

#: 글 채널에 실을 최소 글자수. 이보다 짧은 조각은 쪽번호·머리말이라 검색을 흐린다.
#: 화면의 채널별 목록도 이 기준을 그대로 쓴다 — 두 곳이 다르면 개수가 어긋난다.
MIN_TEXT_CHARS = 20


class ParseError(RuntimeError):
    """파싱을 시작조차 할 수 없는 상태. 조용히 빈 결과를 주지 않는다."""


@dataclass
class TextBlock:
    page: int
    idx: int
    text: str
    char_len: int = 0

    def __post_init__(self) -> None:
        self.char_len = len(self.text)

    @property
    def anchor(self) -> str:
        return f"p{self.page}/t{self.idx}"


@dataclass
class TableBlock:
    page: int
    idx: int
    header: list[str]
    rows: list[list[str]]
    n_numeric_cells: int = 0
    caption: str = ""

    @property
    def anchor(self) -> str:
        return f"p{self.page}/tbl{self.idx}"

    @property
    def shape(self) -> tuple[int, int]:
        return (len(self.rows), len(self.header) if self.header else 0)

    def cell_anchor(self, row: int, col: int) -> str:
        return f"{self.anchor}/r{row}c{col}"


@dataclass
class ImageBlock:
    page: int
    idx: int
    width: int
    height: int
    kind: str = "photo"
    nearby_caption: str = ""

    @property
    def anchor(self) -> str:
        return f"p{self.page}/img{self.idx}"


@dataclass
class ParsedDocument:
    filename: str
    doc_hash: str
    n_pages: int
    text_blocks: list[TextBlock] = field(default_factory=list)
    tables: list[TableBlock] = field(default_factory=list)
    images: list[ImageBlock] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    # ---- 파생 통계 ------------------------------------------------------- #
    @property
    def full_text(self) -> str:
        return "\n".join(b.text for b in self.text_blocks)

    @property
    def table_text(self) -> str:
        """표의 셀 텍스트만 이어 붙인 것. 분류·지표 탐지가 본문과 함께 쓴다 —
        설비명과 지표명은 본문보다 표에 더 정확하게 적혀 있다."""
        parts: list[str] = []
        for t in self.tables:
            parts.append(" ".join(t.header))
            parts.extend(" ".join(r) for r in t.rows)
        return "\n".join(parts)

    @property
    def searchable_text(self) -> str:
        return self.full_text + "\n" + self.table_text

    @property
    def n_numeric_cells(self) -> int:
        return sum(t.n_numeric_cells for t in self.tables)

    def summary(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "doc_hash": self.doc_hash,
            "pages": self.n_pages,
            "text_blocks": len(self.text_blocks),
            "text_chars": sum(b.char_len for b in self.text_blocks),
            "tables": len(self.tables),
            "table_rows": sum(len(t.rows) for t in self.tables),
            "numeric_cells": self.n_numeric_cells,
            "images": len(self.images),
            "image_kinds": _count(i.kind for i in self.images),
            "warnings": list(self.warnings),
        }


def _count(it) -> dict[str, int]:
    out: dict[str, int] = {}
    for v in it:
        out[v] = out.get(v, 0) + 1
    return dict(sorted(out.items()))


def _clean(cell: Any) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell)).strip()


def _is_numeric(s: str) -> bool:
    if not s:
        return False
    return bool(NUMERIC.fullmatch(s.replace(" ", "")))


def _classify_image(width: int, height: int, page_text: str) -> str:
    """이미지 종류 추정. 확정이 아니라 제안 — 사람이 뒤집을 수 있다."""
    if width < MIN_IMAGE_PX or height < MIN_IMAGE_PX:
        return "logo"
    ratio = width / max(height, 1)
    txt = page_text or ""
    if any(k in txt for k in ("차트", "추이", "그래프")):
        return "chart"
    if any(k in txt for k in ("도면", "배치도", "구조도", "외형도", "제작도")):
        return "drawing"
    if ratio > 3.0 or ratio < 0.33:
        return "drawing"
    return "photo"


def parse_pdf(path: str, *, extract_images: bool = True) -> ParsedDocument:
    """PDF 를 4채널로 분해한다.

    표 추출은 pdfplumber 의 격자 인식을 쓴다. 격자선이 없는 표(선 없는 레이아웃)는
    놓칠 수 있고, 그 경우 **경고를 남긴다** — 조용히 빈 결과를 주면 표가 없는
    문서인지 파서가 못 읽은 것인지 구분할 수 없다.
    """
    try:
        import pdfplumber
    except ImportError as exc:  # pragma: no cover - 설치 안내 경로
        raise ParseError("pdfplumber 가 필요하다 (uv pip install pdfplumber)") from exc

    with open(path, "rb") as f:
        raw = f.read()
    doc_hash = hashlib.sha256(raw).hexdigest()[:16]

    doc = ParsedDocument(filename=os.path.basename(path), doc_hash=doc_hash, n_pages=0)

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        doc.n_pages = len(pdf.pages)
        for pno, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""

            # --- 글 ------------------------------------------------------- #
            if page_text.strip():
                for bidx, chunk in enumerate(_split_paragraphs(page_text)):
                    doc.text_blocks.append(TextBlock(page=pno, idx=bidx, text=chunk))

            # --- 표 ------------------------------------------------------- #
            try:
                raw_tables = page.extract_tables()
            except Exception as exc:  # noqa: BLE001 - 한 면이 깨져도 나머지는 읽는다
                raw_tables = []
                doc.warnings.append(f"p{pno} 표 추출 실패: {exc}")

            tidx = 0
            for t in raw_tables:
                grid = [[_clean(c) for c in row] for row in t]
                grid = [r for r in grid if any(c for c in r)]
                if len(grid) < 2:
                    continue  # 사진 캡션용 1행짜리 격자는 표가 아니다
                header, rows = grid[0], grid[1:]
                n_num = sum(1 for r in rows for c in r if _is_numeric(c))
                if n_num == 0 and len(rows) < 2:
                    continue
                doc.tables.append(
                    TableBlock(
                        page=pno, idx=tidx, header=header, rows=rows,
                        n_numeric_cells=n_num, caption=_guess_caption(page_text),
                    )
                )
                tidx += 1

            # --- 그림 ----------------------------------------------------- #
            if extract_images:
                for iidx, im in enumerate(page.images or []):
                    w = int(abs(im.get("x1", 0) - im.get("x0", 0)))
                    h = int(abs(im.get("bottom", 0) - im.get("top", 0)))
                    doc.images.append(
                        ImageBlock(
                            page=pno, idx=iidx, width=w, height=h,
                            kind=_classify_image(w, h, page_text),
                            nearby_caption=_guess_caption(page_text),
                        )
                    )

    if not doc.tables:
        doc.warnings.append(
            "표가 하나도 추출되지 않았다. 격자선 없는 레이아웃이거나 스캔본일 수 있다 "
            "(v0.2 OCR 대상)."
        )
    if not doc.text_blocks:
        doc.warnings.append(
            "텍스트 레이어가 없다. 스캔본 PDF 로 보인다 — 지금 파서로는 빈 결과가 된다."
        )
    return doc


def _split_paragraphs(text: str) -> list[str]:
    parts = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    return parts or ([text.strip()] if text.strip() else [])


CAPTION = re.compile(r"\[\s*([^\]]{2,40})\s*\]")


def _guess_caption(page_text: str) -> str:
    m = CAPTION.search(page_text or "")
    return m.group(1).strip() if m else ""


# --------------------------------------------------------------------------- #
# 엑셀 채널 — 표를 그대로 시트로 떨군다.
# `llmwiki/server/excel.py` 가 명세서를 내보내는 것과 같은 역할이다.
# --------------------------------------------------------------------------- #
def to_excel(doc: ParsedDocument, out_path: str) -> str:
    """추출한 표 전체를 xlsx 로. 시트 1장 = 표 1개, 첫 시트는 목차."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover - 설치 안내 경로
        raise ParseError("openpyxl 이 필요하다 (uv pip install openpyxl)") from exc

    wb = Workbook()
    idx_ws = wb.active
    idx_ws.title = "목차"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4F46E5")

    idx_ws.append(["시트", "페이지", "표", "행", "열", "숫자셀", "캡션"])
    for c in idx_ws[1]:
        c.font, c.fill = hdr_font, hdr_fill

    for t in doc.tables:
        sheet = f"p{t.page}_t{t.idx}"[:31]
        ws = wb.create_sheet(sheet)
        if t.header:
            ws.append(t.header)
            for c in ws[1]:
                c.font, c.fill = hdr_font, hdr_fill
                c.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in t.rows:
            ws.append(row)
        _fit_columns(ws, minimum=8)
        r, c = t.shape
        idx_ws.append([sheet, t.page, t.idx, r, c, t.n_numeric_cells, t.caption])

    _fit_columns(idx_ws, minimum=10)
    wb.save(out_path)
    return out_path


def _fit_columns(ws, *, minimum: int) -> None:
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=minimum)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, minimum), 40)


# --------------------------------------------------------------------------- #
# 검색 인덱싱용 청크 — 채널마다 다르게 만든다.
# --------------------------------------------------------------------------- #
def to_chunks(doc: ParsedDocument) -> list[dict]:
    """검색 단위. **표는 표 단위로 통째 유지한다.**

    표를 문장처럼 잘라 넣으면 행-열 관계가 다시 깨진다. 표 1개 = 청크 1개로 두고
    마크다운 파이프 형식으로 직렬화해, 검색된 컨텍스트 안에서도 격자가 살아 있게 한다.
    """
    chunks: list[dict] = []

    for b in doc.text_blocks:
        if b.char_len < MIN_TEXT_CHARS:
            continue
        chunks.append({
            "channel": "text",
            "anchor": b.anchor,
            "page": b.page,
            "content": b.text,
        })

    for t in doc.tables:
        lines = [f"[표: {t.caption}] (p.{t.page})" if t.caption else f"[표] (p.{t.page})"]
        if t.header:
            lines.append(" | ".join(t.header))
            lines.append(" | ".join("---" for _ in t.header))
        lines.extend(" | ".join(row) for row in t.rows)
        chunks.append({
            "channel": "table",
            "anchor": t.anchor,
            "page": t.page,
            "content": "\n".join(lines),
            "n_numeric_cells": t.n_numeric_cells,
        })

    for im in doc.images:
        if im.kind == "logo":
            continue
        cap = im.nearby_caption or "캡션 없음"
        chunks.append({
            "channel": "image",
            "anchor": im.anchor,
            "page": im.page,
            "content": f"[그림/{im.kind}] p.{im.page} — {cap} ({im.width}x{im.height})",
        })

    return chunks


def to_dict(doc: ParsedDocument) -> dict:
    return {
        "summary": doc.summary(),
        "text_blocks": [asdict(b) for b in doc.text_blocks],
        "tables": [asdict(t) for t in doc.tables],
        "images": [asdict(i) for i in doc.images],
    }
