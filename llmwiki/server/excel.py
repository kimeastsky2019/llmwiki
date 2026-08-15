"""프로그램 명세서 → Excel 산출물."""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..i18n import DOC_LABELS, normalize, xlsx

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
WRAP = Alignment(vertical="top", wrap_text=True)


def build_workbook(meta: dict[str, Any], body: str, impact: dict[str, Any],
                   statements: list[dict[str, Any]], lang: str = "ko") -> bytes:
    lang = normalize(lang)
    wb = Workbook()

    _sheet_overview(wb.active, meta, body, lang)
    _sheet_list(wb.create_sheet(xlsx("sheet_classes", lang)), xlsx("title_classes", lang),
                xlsx("head_class_fqn", lang), [[c] for c in meta.get("classes", [])])
    _sheet_crud(wb.create_sheet(xlsx("sheet_crud", lang)), body, meta, lang)
    _sheet_sql(wb.create_sheet(xlsx("sheet_sql", lang)), statements, lang)
    _sheet_impact(wb.create_sheet(xlsx("sheet_impact", lang)), impact, lang)
    _sheet_list(wb.create_sheet(xlsx("sheet_source", lang)), xlsx("title_source", lang),
                xlsx("head_source", lang), [[f] for f in meta.get("files", [])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_overview(ws, meta: dict[str, Any], body: str, lang: str) -> None:
    ws.title = xlsx("sheet_overview", lang)
    ws["A1"] = meta.get("name", meta.get("id", ""))
    ws["A1"].font = TITLE_FONT
    values = [
        meta.get("id", ""),
        meta.get("name", ""),
        meta.get("layer", ""),
        meta.get("entry", ""),
        ", ".join(meta.get("urls", []) or []),
        ", ".join(meta.get("service_ids", []) or []),
        ", ".join(meta.get("tables", []) or []),
        len(meta.get("sql_ids", []) or []),
        meta.get("generated_at", ""),
        meta.get("generator", ""),
    ]
    rows = list(zip(xlsx("overview_rows", lang), values))
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v)).alignment = WRAP

    start = len(rows) + 5
    ws.cell(row=start, column=1, value=xlsx("body_title", lang)).font = Font(bold=True)
    for i, line in enumerate(body.splitlines(), start=start + 1):
        ws.cell(row=i, column=1, value=line).alignment = WRAP

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 90


def _sheet_list(ws, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for i, row in enumerate(rows, start=4):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value).alignment = WRAP
    _autosize(ws, len(headers))


def _sheet_crud(ws, body: str, meta: dict[str, Any], lang: str) -> None:
    rows = _parse_crud(body)
    if not rows:
        rows = [[t, "", "", "", ""] for t in meta.get("tables", [])]
    _sheet_list(ws, xlsx("title_crud", lang), xlsx("head_crud", lang), rows)


def _sheet_sql(ws, statements: list[dict[str, Any]], lang: str) -> None:
    rows = [
        [s["full_id"], s["kind"], ", ".join(s.get("params", [])), s.get("path", ""),
         s.get("line", ""), s.get("sql", "")]
        for s in statements
    ]
    _sheet_list(ws, xlsx("title_sql", lang), xlsx("head_sql", lang), rows)
    ws.column_dimensions[get_column_letter(6)].width = 100


def _sheet_impact(ws, impact: dict[str, Any], lang: str) -> None:
    rows = [
        [r["id"], r["name"], r["layer"], ", ".join(r["tables"])]
        for r in impact.get("affected", [])
    ]
    _sheet_list(ws, xlsx("title_impact", lang), xlsx("head_impact", lang), rows)


# 부록 B 의 제목·헤더는 언어마다 다르므로 두 언어 모두 인식한다
# (문서의 language 와 내려받기 언어가 어긋나도 표를 놓치지 않도록).
_CRUD_HEADINGS = {DOC_LABELS[lg]["appendix_b"] for lg in DOC_LABELS}
_CRUD_FIRST_CELLS = {DOC_LABELS[lg]["crud_table_col"] for lg in DOC_LABELS}


def _parse_crud(body: str) -> list[list[str]]:
    out: list[list[str]] = []
    in_table = False
    for line in body.splitlines():
        if line.startswith("## "):
            heading = line[3:].strip()
            if heading in _CRUD_HEADINGS:
                in_table = True
                continue
            if in_table:
                break
        if not in_table or not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) != 5 or cells[0] in _CRUD_FIRST_CELLS:
            continue
        # 구분선(|---|:-:|…) 만 걸러낸다. 빈 셀은 'CRUD 해당 없음' 이지 구분선이 아니다.
        if all(c and set(c) <= {"-", ":"} for c in cells):
            continue
        out.append(cells)
    return out


def _autosize(ws, ncols: int) -> None:
    for j in range(1, ncols + 1):
        letter = get_column_letter(j)
        width = max(
            (len(str(ws.cell(row=i, column=j).value or "")) for i in range(3, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip() or "program"
