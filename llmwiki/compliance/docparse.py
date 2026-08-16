"""문서 파서 (L0) — docx · xlsx · pdf → 평문 + 절 + 문자 오프셋.

여기서 나오는 `text` 가 **모든 근거 스팬의 기준 좌표계**다. 파서가 같은 파일에서
매번 같은 문자열을 만들어야 어제 붙인 근거가 오늘도 같은 문장을 가리킨다.
그래서 이 파일에는 "보기 좋게" 하는 처리가 없다 — 줄바꿈 하나까지 규칙으로 고정한다.

docx 에서 제일 어려운 것: **조문 번호가 문서에 없다**
--------------------------------------------------
실제 규정 문서(AI거버넌스규정.docx)를 열어 보면 본문에 "제1조" 라는 글자가 없다.
Word 가 `Heading3` 스타일에 걸린 번호매기기 정의(`제%3조`)로 화면에만 그려 준다.
그대로 텍스트만 뽑으면 조문 번호가 통째로 사라지고, 조문 앵커를 만들 수 없다.

그래서 styles.xml → numbering.xml 을 따라가 번호를 **재구성**한다. 문서 순서대로
레벨별 카운터를 돌려 `제1조`, `제2조` … 를 만들어 문단 앞에 붙인다. 이걸 하지 않으면
이 시스템은 실제 규정 문서에 아예 쓸 수 없다.

표를 텍스트로 펴는 규칙
----------------------
행마다 `| 셀 | 셀 |` 한 줄. 업무 문서의 내용은 대부분 표 안에 있어서(구분/내용 표),
표를 버리면 근거로 쓸 문장이 남지 않는다. 셀 안의 줄바꿈은 공백으로 접는다 —
그래야 한 셀이 한 근거 단위가 된다.
"""

from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

SUPPORTED = (".docx", ".xlsx", ".pdf", ".txt", ".md")


# --------------------------------------------------------------------------- #
# 결과 모델
# --------------------------------------------------------------------------- #
@dataclass
class Section:
    """문서의 한 절. start/end 는 ParsedDoc.text 의 문자 오프셋이다."""

    level: int
    number: str
    title: str
    start: int
    end: int
    path: str = ""
    #: 번호만으로 만든 경로 (`제2장/제2절/제1조`).
    #:
    #: ★ 실제 규정 문서에서 조문 번호는 절마다 1부터 다시 시작한다.
    #:   AI거버넌스규정.docx 에는 "제1조" 가 **14개** 있다. 번호 하나로는 조문을
    #:   가리킬 수 없고, 이것을 식별자로 쓰면 매핑이 통째로 어긋난다.
    #:   앵커는 이 경로에서 유도한다. 제목은 개정 때 문구가 바뀌므로 빼 둔다.
    number_path: str = ""

    @property
    def label(self) -> str:
        return f"{self.number} {self.title}".strip()

    def to_dict(self) -> dict[str, Any]:
        return {
            "level": self.level, "number": self.number, "title": self.title,
            "start": self.start, "end": self.end, "path": self.path,
            "number_path": self.number_path,
        }


@dataclass
class ParsedDoc:
    doc_id: str
    source: str
    kind: str
    text: str
    sections: list[Section] = field(default_factory=list)
    tables: list[list[list[str]]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def sha256(self) -> str:
        return hashlib.sha256(self.text.encode("utf-8")).hexdigest()

    def body_of(self, section: Section) -> str:
        return self.text[section.start:section.end]

    def section_labels(self) -> list[str]:
        return [s.label for s in self.sections]

    def find(self, needle: str) -> Section | None:
        flat = _flat(needle)
        for s in self.sections:
            if flat and flat in _flat(s.label):
                return s
        return None

    def to_dict(self) -> dict[str, Any]:
        return {
            "doc_id": self.doc_id, "source": self.source, "kind": self.kind,
            "chars": len(self.text), "sha256": self.sha256[:16],
            "sections": [s.to_dict() for s in self.sections],
            "warnings": self.warnings,
        }


class ParseError(RuntimeError):
    pass


# --------------------------------------------------------------------------- #
# 진입점
# --------------------------------------------------------------------------- #
def parse(path: str | Path, *, doc_id: str | None = None) -> ParsedDoc:
    p = Path(path)
    if not p.exists():
        raise ParseError(f"파일이 없다: {p}")
    ident = doc_id or slug(p.stem)
    suffix = p.suffix.lower()
    if suffix == ".docx":
        return parse_docx(p, ident)
    if suffix == ".xlsx":
        return parse_xlsx(p, ident)
    if suffix == ".pdf":
        return parse_pdf(p, ident)
    if suffix in (".txt", ".md"):
        text = p.read_text(encoding="utf-8", errors="replace")
        doc = ParsedDoc(ident, p.name, suffix.lstrip("."), text)
        doc.sections = detect_sections(text)
        return doc
    raise ParseError(
        f"지원하지 않는 형식: {p.suffix} (지원: {', '.join(SUPPORTED)})"
    )


def slug(name: str) -> str:
    out = re.sub(r"[^0-9A-Za-z가-힣]+", "-", name).strip("-").lower()
    return out[:80] or "doc"


# --------------------------------------------------------------------------- #
# docx
# --------------------------------------------------------------------------- #
def parse_docx(path: Path, doc_id: str) -> ParsedDoc:
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        if "word/document.xml" not in names:
            raise ParseError(f"docx 로 보이지 않는다: {path.name}")
        document = ET.fromstring(z.read("word/document.xml"))
        numbering = (
            ET.fromstring(z.read("word/numbering.xml"))
            if "word/numbering.xml" in names else None
        )
        styles = (
            ET.fromstring(z.read("word/styles.xml"))
            if "word/styles.xml" in names else None
        )

    numberer = _Numberer(numbering, styles)
    body = document.find(f"{W}body")
    lines: list[str] = []
    tables: list[list[list[str]]] = []
    # (레벨, 번호, 제목, 줄 인덱스) — 오프셋은 본문을 다 만든 뒤에 계산한다
    marks: list[tuple[int, str, str, int]] = []

    for el in list(body) if body is not None else []:
        tag = el.tag
        if tag == f"{W}p":
            text = _para_text(el)
            number, level = numberer.number_for(el)
            if not text.strip() and not number:
                continue
            line = f"{number} {text}".strip() if number else text
            # 절로 잡는 것은 개요 수준을 가진 것(Heading 스타일)과 장·절·조뿐이다.
            # 각 호(1. 2. 가.)까지 절로 만들면 조문 구조가 그 안에 묻힌다 —
            # 호는 조의 '내용'이지 조와 나란한 단위가 아니다.
            if level is not None and _looks_like_heading(text):
                marks.append((level, number, _strip_leading(text), len(lines)))
            elif _STRUCTURAL_RE.match(number):
                marks.append((_structural_level(number), number,
                              _strip_leading(text), len(lines)))
            lines.append(line)
        elif tag == f"{W}tbl":
            rows = _table_rows(el)
            if rows:
                tables.append(rows)
                lines.extend(_render_table(rows))

    text = "\n".join(lines) + "\n"
    doc = ParsedDoc(doc_id, path.name, "docx", text, tables=tables)
    # 개요 스타일이 있는 문서(규정·지침)는 그것을 믿고, 없는 문서(별첨 서식)는
    # 본문 머리표로 찾는다. 두 종류가 실제로 섞여 들어온다.
    doc.sections = _sections_from_marks(text, lines, marks) if len(marks) >= 3 else []
    if not doc.sections:
        doc.sections = detect_sections(text)
    if numbering is None:
        doc.warnings.append("numbering.xml 이 없어 자동 번호를 복원하지 못했다")
    return doc


#: 장·절·조는 개요 스타일이 없어도 문서의 뼈대다.
_STRUCTURAL_RE = re.compile(r"^제\s*\d+\s*(장|절|조)")


def _structural_level(number: str) -> int:
    if "장" in number:
        return 0
    if "절" in number:
        return 1
    return 2


def _para_text(p: ET.Element) -> str:
    """문단 텍스트. 탭·줄바꿈은 공백으로 접어 한 문단을 한 줄로 만든다."""
    out: list[str] = []
    for node in p.iter():
        if node.tag == f"{W}t":
            out.append(node.text or "")
        elif node.tag in (f"{W}tab", f"{W}br"):
            out.append(" ")
    return re.sub(r"[ \t]+", " ", "".join(out)).strip()


def _table_rows(tbl: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in tbl.findall(f"{W}tr"):
        cells: list[str] = []
        for tc in tr.findall(f"{W}tc"):
            parts = [_para_text(p) for p in tc.findall(f"{W}p")]
            cells.append(" ".join(x for x in parts if x).strip())
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _render_table(rows: list[list[str]]) -> list[str]:
    """표 → `| 셀 | 셀 |` 줄. 셀 하나가 근거 한 단위가 되도록 한 행을 한 줄에 둔다."""
    return ["| " + " | ".join(c.replace("|", "/") for c in row) + " |" for row in rows]


class _Numberer:
    """styles.xml + numbering.xml 로 자동 번호를 재구성한다.

    Word 는 번호를 문서에 저장하지 않고 화면에 그린다. 문서 순서대로 레벨별
    카운터를 돌려 같은 문자열을 만들어 낸다 — 여기가 틀리면 조문 번호가 어긋나
    모든 매핑이 깨진다.
    """

    def __init__(self, numbering: ET.Element | None, styles: ET.Element | None) -> None:
        self.levels: dict[tuple[str, int], dict[str, str]] = {}
        self.num_to_abs: dict[str, str] = {}
        self.style_num: dict[str, tuple[str, int]] = {}
        self.counters: dict[tuple[str, int], int] = {}

        if numbering is not None:
            for num in numbering.iter(f"{W}num"):
                target = num.find(f"{W}abstractNumId")
                if target is not None:
                    self.num_to_abs[num.get(f"{W}numId", "")] = target.get(f"{W}val", "")
            for an in numbering.iter(f"{W}abstractNum"):
                aid = an.get(f"{W}abstractNumId", "")
                for lvl in an.findall(f"{W}lvl"):
                    ilvl = int(lvl.get(f"{W}ilvl", "0"))
                    self.levels[(aid, ilvl)] = {
                        "fmt": _val(lvl.find(f"{W}numFmt"), "decimal"),
                        "text": _val(lvl.find(f"{W}lvlText"), ""),
                        "start": _val(lvl.find(f"{W}start"), "1"),
                    }
        if styles is not None:
            for style in styles.iter(f"{W}style"):
                num = style.find(f"{W}pPr/{W}numPr")
                if num is None:
                    continue
                nid = _val(num.find(f"{W}numId"), "")
                ilvl = _val(num.find(f"{W}ilvl"), "")
                sid = style.get(f"{W}styleId", "")
                if nid:
                    self.style_num[sid] = (nid, int(ilvl) if ilvl.isdigit() else _heading_level(sid))

    def number_for(self, p: ET.Element) -> tuple[str, int | None]:
        """(번호 문자열, 개요 수준). 번호가 없으면 ("", 수준 또는 None)."""
        style = _val(p.find(f"{W}pPr/{W}pStyle"), "")
        direct = p.find(f"{W}pPr/{W}numPr")
        nid, ilvl = "", 0
        if direct is not None:
            nid = _val(direct.find(f"{W}numId"), "")
            raw = _val(direct.find(f"{W}ilvl"), "0")
            ilvl = int(raw) if raw.isdigit() else 0
        elif style in self.style_num:
            nid, ilvl = self.style_num[style]

        outline = _heading_level(style) if style.startswith("Heading") else None
        if not nid or nid == "0":
            return "", outline

        aid = self.num_to_abs.get(nid, "")
        spec = self.levels.get((aid, ilvl))
        if spec is None or spec["fmt"] in ("none", "bullet") or not spec["text"]:
            return "", outline

        key = (aid, ilvl)
        self.counters[key] = self.counters.get(key, int(spec["start"]) - 1) + 1
        # 하위 레벨은 다시 1부터 — 제2조로 넘어가면 그 아래 항 번호가 이어지면 안 된다
        for (a, l) in list(self.counters):
            if a == aid and l > ilvl:
                del self.counters[(a, l)]

        rendered = spec["text"]
        for depth in range(9):
            token = f"%{depth + 1}"
            if token not in rendered:
                continue
            value = self.counters.get((aid, depth), int(
                self.levels.get((aid, depth), {}).get("start", "1")))
            fmt = self.levels.get((aid, depth), {}).get("fmt", "decimal")
            rendered = rendered.replace(token, _format_number(value, fmt))
        # 목록 깊이(ilvl)를 개요 수준으로 흘려보내지 않는다. 각 호의 ilvl 은 0 이라
        # 그대로 쓰면 '1.' 이 '제1조'(Heading3 = 2) 보다 위 단계로 잡힌다.
        return rendered.strip(), outline


_GANADA = "가나다라마바사아자차카타파하"
_ROMAN = [
    (1000, "m"), (900, "cm"), (500, "d"), (400, "cd"), (100, "c"), (90, "xc"),
    (50, "l"), (40, "xl"), (10, "x"), (9, "ix"), (5, "v"), (4, "iv"), (1, "i"),
]


def _format_number(value: int, fmt: str) -> str:
    if value < 1:
        value = 1
    if fmt == "ganada":
        return _GANADA[(value - 1) % len(_GANADA)]
    if fmt == "upperLetter":
        return chr(ord("A") + (value - 1) % 26)
    if fmt == "lowerLetter":
        return chr(ord("a") + (value - 1) % 26)
    if fmt in ("upperRoman", "lowerRoman"):
        out, n = "", value
        for weight, sign in _ROMAN:
            while n >= weight:
                out += sign
                n -= weight
        return out.upper() if fmt == "upperRoman" else out
    # decimalFullWidth 는 전각 숫자지만 반각으로 적는다 —
    # 사람이 검색하고 매핑에 쓰는 문자열이라 '제１조' 보다 '제1조' 가 맞다.
    return str(value)


def _heading_level(style_id: str) -> int:
    m = re.search(r"(\d)", style_id)
    return int(m.group(1)) - 1 if m else 0


def _val(node: ET.Element | None, default: str) -> str:
    if node is None:
        return default
    return node.get(f"{W}val", default)


def _looks_like_heading(text: str) -> bool:
    return bool(text.strip()) and len(text) < 200


def _strip_leading(text: str) -> str:
    return re.sub(r"^[\s.·]+", "", text).strip()


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #
def parse_xlsx(path: Path, doc_id: str) -> ParsedDoc:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError("openpyxl 이 필요하다") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    tables: list[list[list[str]]] = []
    marks: list[tuple[int, str, str, int]] = []
    try:
        for sheet in wb.worksheets:
            marks.append((0, "[시트]", str(sheet.title), len(lines)))
            lines.append(f"# {sheet.title}")
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).replace("\n", " ").strip() for v in row]
                while cells and not cells[-1]:
                    cells.pop()
                if not any(cells):
                    continue
                rows.append(cells)
                lines.append("| " + " | ".join(c.replace("|", "/") for c in cells) + " |")
            if rows:
                tables.append(rows)
    finally:
        wb.close()

    text = "\n".join(lines) + "\n"
    doc = ParsedDoc(doc_id, path.name, "xlsx", text, tables=tables)
    doc.sections = _sections_from_marks(text, lines, marks)
    return doc


# --------------------------------------------------------------------------- #
# pdf
# --------------------------------------------------------------------------- #
def parse_pdf(path: Path, doc_id: str) -> ParsedDoc:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise ParseError("pypdf 가 필요하다 (uv pip install pypdf)") from exc

    reader = PdfReader(str(path))
    lines: list[str] = []
    marks: list[tuple[int, str, str, int]] = []
    warnings: list[str] = []
    empty = 0
    for i, page in enumerate(reader.pages, start=1):
        marks.append((0, f"p.{i}", "", len(lines)))
        body = (page.extract_text() or "").strip()
        if not body:
            empty += 1
        for raw in body.splitlines():
            lines.append(re.sub(r"[ \t]+", " ", raw).rstrip())
    text = "\n".join(lines) + "\n"
    if empty:
        # 스캔 PDF 는 글자가 없다. 근거 스팬을 만들 수 없으므로 조용히 넘기면 안 된다.
        warnings.append(
            f"글자를 뽑지 못한 쪽이 {empty}쪽 있다 — 스캔본이면 OCR 이 필요하다"
        )
    doc = ParsedDoc(doc_id, path.name, "pdf", text, warnings=warnings)
    doc.sections = detect_sections(text) or _sections_from_marks(text, lines, marks)
    return doc


# --------------------------------------------------------------------------- #
# 절 인식 — 스타일이 없는 문서를 위한 규칙
# --------------------------------------------------------------------------- #
#: 업무 문서에서 실제로 쓰이는 머리표. 순서가 곧 우선순위다.
_HEAD_PATTERNS: tuple[tuple[str, int], ...] = (
    (r"제\s*(\d+)\s*장", 0),
    (r"제\s*(\d+)\s*절", 1),
    (r"제\s*(\d+)\s*조(?:\s*의\s*\d+)?", 2),
    # 영문 규정. 한국어 조문과 같은 수준으로 잡아 같은 경로를 타게 한다.
    (r"Chapter\s+(\d+)", 0),
    (r"Section\s+(\d+)", 1),
    (r"Article\s+(\d+)", 2),
    (r"([IVXivx]+)\.", 1),
    (r"(붙임|별첨|부록)\s*\.?\s*\d*", 1),
    (r"(Annex|Appendix|Attachment)\s*\.?\s*\d*", 1),
    (r"(\d+)\.", 2),
    (r"([가-힣])\.", 3),
)

_HEAD_RE = re.compile(
    r"^\s*(?:"
    + "|".join(f"(?P<g{i}>{p})" for i, (p, _) in enumerate(_HEAD_PATTERNS))
    + r")\s*(?P<rest>.*)$"
)


def _pattern_heading(line: str) -> tuple[int, str, str] | None:
    """스타일이 없는 문서에서 머리글줄을 찾는다. (수준, 번호, 제목)"""
    stripped = line.strip()
    if not stripped or len(stripped) > 120 or stripped.startswith("|"):
        return None
    m = _HEAD_RE.match(stripped)
    if not m:
        return None
    for i, (_, level) in enumerate(_HEAD_PATTERNS):
        matched = m.group(f"g{i}")
        if matched:
            rest = m.group("rest").strip()
            # "1. " 뒤에 아무것도 없으면 목록 항목이지 절이 아니다
            if not rest:
                return None
            return level, matched.strip(), _strip_leading(rest)
    return None


def detect_sections(text: str) -> list[Section]:
    """평문에서 절을 찾는다 (docx 스타일 정보가 없을 때의 경로)."""
    lines = text.split("\n")
    marks: list[tuple[int, str, str, int]] = []
    for i, line in enumerate(lines):
        found = _pattern_heading(line)
        if found:
            marks.append((found[0], found[1], found[2], i))
    return _sections_from_marks(text, lines, marks)


def _sections_from_marks(
    text: str, lines: list[str], marks: list[tuple[int, str, str, int]]
) -> list[Section]:
    """줄 인덱스로 잡은 머리글을 문자 오프셋 구간으로 바꾼다."""
    if not marks:
        return []
    # 줄 시작 오프셋 (join("\n") 규칙과 정확히 같아야 한다)
    starts: list[int] = []
    pos = 0
    for line in lines:
        starts.append(pos)
        pos += len(line) + 1

    sections: list[Section] = []
    for idx, (level, number, title, line_no) in enumerate(marks):
        start = starts[line_no]
        end = starts[marks[idx + 1][3]] if idx + 1 < len(marks) else len(text)
        sections.append(Section(level, number, title, start, end))

    # 상위 절 경로. 사람이 읽는 path 와 앵커용 number_path 를 함께 만든다.
    trail: dict[int, str] = {}
    numbers: dict[int, str] = {}
    for s in sections:
        trail = {k: v for k, v in trail.items() if k < s.level}
        numbers = {k: v for k, v in numbers.items() if k < s.level}
        s.path = " > ".join([*(trail[k] for k in sorted(trail)), s.label])
        s.number_path = "/".join(
            [*(numbers[k] for k in sorted(numbers)), s.number or s.title]
        )
        trail[s.level] = s.label
        numbers[s.level] = s.number or s.title
    return sections


# --------------------------------------------------------------------------- #
def _flat(text: str) -> str:
    return re.sub(r"\s+", "", text)


def walk(root: str | Path, *, suffixes: Iterable[str] = SUPPORTED) -> list[Path]:
    """폴더에서 파싱 가능한 문서를 모은다. 임시 파일(~$…)은 건너뛴다."""
    allowed = {s.lower() for s in suffixes}
    out: list[Path] = []
    for p in sorted(Path(root).rglob("*")):
        if p.is_file() and p.suffix.lower() in allowed and not p.name.startswith("~$"):
            out.append(p)
    return out
