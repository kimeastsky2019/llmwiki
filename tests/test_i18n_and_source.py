"""다국어(ko/en) 산출물과 소스 브라우저 API 회귀 테스트."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from llmwiki.config import load_config
from llmwiki.docgen.generator import _crud_rows, _statement_dict
from llmwiki.docgen.prompts import build_prompt, system_prompt
from llmwiki.indexer import scan
from llmwiki.llm.template import TemplateProvider
from llmwiki.server.excel import _parse_crud, build_workbook

ROOT = Path(__file__).resolve().parents[1]

# 활성 프로젝트는 사용자가 뷰어에서 바꿀 수 있으므로 테스트는 항상 명시한다
DEFAULT = {"project": "default"}


@pytest.fixture(scope="module")
def index():
    return scan(load_config(ROOT / "config.yaml"))


@pytest.fixture(scope="module")
def client():
    from llmwiki.server.app import app

    return TestClient(app)


# --------------------------------------------------------------------------- #
# 프롬프트 · 템플릿 공급자
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("lang", ["ko", "en"])
def test_system_prompt_language(lang):
    assert "MyBatis" in system_prompt(lang)


def _render_template(index, lang: str) -> str:
    prog = index.programs[0]
    prompt = build_prompt(
        program_name=prog.name,
        layer=prog.layer,
        entry_fqn=prog.entry_fqn,
        urls=prog.urls,
        tables=prog.tables,
        crud_rows=_crud_rows(index, prog),
        sources=[("A.java", "public class A {}")],
        statements=[
            _statement_dict(index.statements[s])
            for s in prog.sql_ids
            if s in index.statements
        ],
        lang=lang,
    )
    return TemplateProvider().complete(system_prompt(lang), prompt)


def test_template_provider_follows_prompt_language(index):
    """template 공급자는 프롬프트에서 언어를 판별한다 — 섞이면 산출물이 반쪽이 된다."""
    ko = _render_template(index, "ko")
    en = _render_template(index, "en")
    assert "## 1. 개요" in ko and "## 1. Overview" not in ko
    assert "## 1. Overview" in en and "## 1. 개요" not in en


def test_template_provider_keeps_facts(index):
    """언어를 바꿔도 파서가 뽑은 사실(테이블·URL)은 그대로 실려야 한다."""
    prog = index.programs[0]
    for lang in ("ko", "en"):
        body = _render_template(index, lang)
        for table in prog.tables:
            assert table in body
        for url in prog.urls:
            assert url in body


def test_template_provider_drops_none_placeholder(index):
    """값이 없을 때 넣는 '(없음)/(none)' 자리표시자를 목록 항목으로 세지 않는다."""
    prompt = build_prompt(
        program_name="X",
        layer="공통",
        entry_fqn="com.x.X",
        urls=[],
        tables=[],
        crud_rows=[],
        sources=[],
        statements=[],
        lang="en",
    )
    body = TemplateProvider().complete("", prompt)
    assert "| `(none)` |" not in body


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    "lang,sheets,table_head",
    [
        ("ko", ["개요", "클래스", "CRUD", "SQL", "영향도", "소스"], "테이블"),
        ("en", ["Overview", "Classes", "CRUD", "SQL", "Impact", "Sources"], "Table"),
    ],
)
def test_workbook_sheets_by_language(lang, sheets, table_head):
    wb = load_workbook(
        io.BytesIO(build_workbook({"name": "X", "tables": ["TB_A"]}, "", {}, [], lang))
    )
    assert wb.sheetnames == sheets
    assert wb[wb.sheetnames[2]].cell(row=3, column=1).value == table_head


@pytest.mark.parametrize(
    "heading,header",
    [("## 부록 B. CRUD 매트릭스", "| 테이블 |"), ("## Appendix B. CRUD Matrix", "| Table |")],
)
def test_parse_crud_reads_both_languages(heading, header):
    body = "\n".join(
        [
            "## 1. 개요",
            heading,
            f"{header} C | R | U | D |",
            "|---|:-:|:-:|:-:|:-:|",
            "| TB_CUST | ● |  | ● |  |",
            "## 부록 C",
            "| ignore | a | b | c | d |",
        ]
    )
    assert _parse_crud(body) == [["TB_CUST", "●", "", "●", ""]]


# --------------------------------------------------------------------------- #
# 소스 브라우저 API
# --------------------------------------------------------------------------- #
def test_source_tree_lists_parsed_files(client):
    roots = client.get("/api/source/tree", params=DEFAULT).json()
    assert roots and roots[0]["name"] == "sample"
    paths = {f["path"]: f for f in roots[0]["files"]}
    controller = "src/main/java/com/gng/inst/cust/CustomerController.java"
    assert controller in paths
    assert paths[controller]["parsed"] is True
    assert paths[controller]["lang"] == "java"


def test_source_returns_content_and_lang(client):
    r = client.get(
        "/api/source",
        params={**DEFAULT, "path": "src/main/java/com/gng/inst/cust/CustomerController.java", "root": 0},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["lang"] == "java"
    assert body["lines"] > 1
    assert "class CustomerController" in body["content"]


@pytest.mark.parametrize(
    "path",
    ["../../../etc/passwd", "../config.yaml", "/etc/passwd"],
)
def test_source_rejects_paths_outside_roots(client, path):
    """source_roots 밖은 어떤 경로 조합으로도 새어 나가면 안 된다."""
    assert client.get("/api/source", params={**DEFAULT, "path": path}).status_code == 404


def test_error_messages_follow_lang_param(client):
    assert client.get("/api/table/NOPE", params={**DEFAULT, "lang": "ko"}).json()["detail"] == (
        "테이블을 찾을 수 없습니다."
    )
    assert client.get("/api/table/NOPE", params={**DEFAULT, "lang": "en"}).json()["detail"] == (
        "Table not found."
    )


def test_meta_exposes_language(client):
    body = client.get("/api/meta", params=DEFAULT).json()
    assert body["language"] in ("ko", "en")
    assert body["source_roots"] == ["sample"]
